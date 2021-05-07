import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

# sheet = workbook.get_sheet_by_name('Sheet1') ... decipricated function
sheet = workbook['Sheet1']
print(type(sheet))

# sheet = workbook.get_sheet_names() # decipricated function
# print(sheet)

cell = sheet['A1'] 
print(cell.value)
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet['C1'].value)
print(str(sheet['C1'].value))

print(sheet.cell(row=1, column=2))
print(sheet['B1'].value)

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

