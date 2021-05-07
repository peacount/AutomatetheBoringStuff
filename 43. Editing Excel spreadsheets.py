import openpyxl
wb = openpyxl.Workbook()

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value == None)
sheet['A1'] = 42
sheet['A2'] = 'Hello'
import os
os.chdir('c:\\Users\\jenny\\Documents')
wb.save('example.xlsx')

sheet2 = wb.create_sheet()
print(wb.get_sheet_names)
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names)
wb.save('example2.xlsx')

wb.create_sheet(index = 0, title = 'My Other Sheet')
wb.save('example3.xlsx')